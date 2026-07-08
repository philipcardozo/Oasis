from pathlib import Path
from openpyxl import load_workbook
from dcf_export import build_dcf_workbook

def check_workbook(path: Path, ticker: str) -> None:
    assert path.exists() and path.stat().st_size > 10_000
    wb = load_workbook(path, data_only=False)
    
    expected_sheets = [
        "Cover",
        f"DCF - Financials.{ticker}",
        f"DCF - Assumptions.{ticker}",
        f"DCF - Assumptions Doc.{ticker}",
        f"DCF - Valuation.{ticker}"
    ]
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"
        
    cover = wb["Cover"]
    assert cover["C13"].value == f"{ticker} DCF", f"Expected Cover title '{ticker} DCF', got '{cover['C13'].value}'"
    assert cover["C21"].value == "Findings"
    
    val = wb[f"DCF - Valuation.{ticker}"]
    assert val["C7"].value == "WACC"
    assert val["C23"].value == "Terminal Growth Rate"
    
    formulas = "\n".join(
        str(cell.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    assert "#REF!" not in formulas

def main() -> None:
    # BLK (BlackRock)
    check_workbook(build_dcf_workbook("BLK", "cash_flow"), "BLK")
    
    # PFE (Pfizer)
    check_workbook(build_dcf_workbook("PFE", "cash_flow"), "PFE")
    
    # MNDY (Monday.com)
    check_workbook(build_dcf_workbook("MNDY", "cash_flow"), "MNDY")
    
    # TSM (TSMC)
    check_workbook(build_dcf_workbook("TSM", "cash_flow"), "TSM")
    
    print("dcf export ok")

if __name__ == "__main__":
    main()
