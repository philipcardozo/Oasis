from __future__ import annotations

import json
import re
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
import urllib.request
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

ROOT = Path(__file__).parent
DATA = ROOT / "graph" / "data"
FACTS = DATA / "companyfacts"
OUTPUTS = ROOT / "outputs" / "dcf"
UNIVERSE = DATA / "universe.json"

USD = '$#,##0;[Red]($#,##0);-'
USD_MM = '$#,##0.0;[Red]($#,##0.0);-'
PER_SHARE = '$0.00;[Red]($0.00);-'
PCT = '0.0%;[Red](0.0%);-'
NUM = '#,##0;[Red](#,##0);-'

# Colors matching user sheet
DARK_BG = "18181A"
LIGHT_GRAY = "E8E8E8"
YELLOW_INPUT = "FFF2CC"
BLUE_TEXT = "0070C0"
RED_TEXT = "FF0000"
GREEN_TEXT = "196B24"
WHITE = "FFFFFF"

# Mapping tickers to domains for logo downloads
TICKER_DOMAINS = {
    "MNDY": "monday.com",
    "NVDA": "nvidia.com",
    "BLK": "blackrock.com",
    "JPM": "jpmorgan.com",
    "PFE": "pfizer.com",
    "WTW": "wtwco.com"
}

TAGS = {
    "revenue": ["RevenueFromContractWithCustomerExcludingAssessedTax", "Revenues", "SalesRevenueNet", "Revenue", "RevenueFromContractsWithCustomers"],
    "ebit": ["OperatingIncomeLoss", "ProfitLossFromOperatingActivities", "ProfitLossFromOperations"],
    "pretax": [
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxes",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments",
        "ProfitLossBeforeTax"
    ],
    "tax": ["IncomeTaxExpenseBenefit", "IncomeTaxExpenseContinuingOperations", "CurrentTaxExpenseIncome"],
    "da": ["DepreciationDepletionAndAmortization", "DepreciationDepletionAndAmortizationExpense", "DepreciationAndAmortization", "Depreciation", "DepreciationExpense"],
    "capex": ["PaymentsToAcquirePropertyPlantAndEquipment", "PaymentsToAcquireProductiveAssets", "PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities"],
    "cfo": ["NetCashProvidedByUsedInOperatingActivities", "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations", "CashFlowsFromUsedInOperatingActivities"],
    "dividends": ["PaymentsOfDividendsCommonStock", "PaymentsOfDividends", "DividendsCommonStockCash", "DividendsPaidOrdinarySharesPerShare"],
    "shares": ["WeightedAverageNumberOfDilutedSharesOutstanding", "WeightedAverageNumberOfSharesOutstandingBasic", "EntityCommonStockSharesOutstanding", "WeightedAverageShares", "AdjustedWeightedAverageShares"],
    "cash": ["CashAndCashEquivalentsAtCarryingValue", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents", "CashAndCashEquivalents", "Cash"],
    "debt": ["LongTermDebtCurrent", "LongTermDebtNoncurrent", "ShortTermBorrowings", "ShortTermDebt", "LongTermDebtAndFinanceLeaseObligationsCurrent", "LongTermDebtAndFinanceLeaseObligationsNoncurrent", "CurrentFinancialLiabilities", "NoncurrentFinancialLiabilities"],
    "st_investments": ["ShortTermInvestments", "CurrentFinancialAssets"],
    "receivables": ["AccountsReceivableNetCurrent", "AccountsReceivableNet", "CurrentTradeReceivables", "TradeAccountReceivable"],
    "current_assets": ["AssetsCurrent", "CurrentAssets"],
    "ppe_net": ["PropertyPlantAndEquipmentNet", "PropertyPlantAndEquipment"],
    "total_assets": ["Assets", "EquityAndLiabilities"],
    "ap": ["AccountsPayableCurrent", "AccountsPayable", "TradeAndOtherCurrentPayablesToTradeSuppliers"],
    "current_liabilities": ["LiabilitiesCurrent", "CurrentLiabilities"],
    "total_liabilities": ["Liabilities", "NoncurrentLiabilities"],
    "equity": ["StockholdersEquity", "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest", "Equity", "EquityAttributableToOwnersOfParent"],
    "common_stock": ["CommonStockValue", "NumberOfSharesIssuedAndFullyPaid"],
    "retained_earnings": ["RetainedEarningsAccumulatedDeficit"],
}

def slug(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")[:48] or "dcf"

def _universe_mtime() -> float:
    return UNIVERSE.stat().st_mtime if UNIVERSE.exists() else 0


@lru_cache(maxsize=1)
def _load_universe_nodes(mtime: float) -> tuple[dict, dict]:
    nodes = json.load(UNIVERSE.open())["nodes"]
    by_id = {n["id"]: n for n in nodes}
    aliases = {str(n.get("t", "")).upper(): n["id"] for n in nodes if n.get("t")}
    aliases.update({n["id"].upper(): n["id"] for n in nodes})
    return by_id, aliases


def load_node(entity_id: str) -> dict:
    by_id, aliases = _load_universe_nodes(_universe_mtime())
    resolved = aliases.get(entity_id.upper(), entity_id)
    if resolved not in by_id:
        raise ValueError(f"unknown entity {entity_id}")
    node = by_id[resolved]
    if node.get("issuer_id") and node["issuer_id"] in by_id:
        return by_id[node["issuer_id"]]
    return node

def load_facts(node: dict) -> tuple[dict, Path]:
    from cache_companyfacts import cache_one, cik10
    cik = cik10(node.get("cik", ""))
    path = FACTS / f"CIK{cik}.json"
    if not path.exists():
        path = cache_one(cik)
    return json.load(path.open()), path

def get_latest_filed_annual(facts: dict, tags: list[str]) -> dict[int, float]:
    picked: dict[int, float] = {}
    picked_filed: dict[int, str] = {}
    for tag in tags:
        for standard in ["us-gaap", "ifrs-full"]:
            rows = facts.get("facts", {}).get(standard, {}).get(tag, {}).get("units", {})
            if not rows:
                continue
            unit_keys = list(rows.keys())
            usd_rows = []
            for uk in ["USD", "shares", "USD/shares", "pure"]:
                if uk in rows:
                    usd_rows = rows[uk]
                    break
            if not usd_rows and unit_keys:
                usd_rows = rows[unit_keys[0]]
            for row in usd_rows:
                if not row.get("fy") or not isinstance(row.get("val"), (int, float)):
                    continue
                annual = row.get("fp") == "FY" or str(row.get("form", "")).upper() in {"10-K", "20-F", "40-F"}
                if annual:
                    fy = int(row["fy"])
                    filed = str(row.get("filed", ""))
                    if fy not in picked_filed or filed > picked_filed[fy]:
                        picked[fy] = float(row["val"])
                        picked_filed[fy] = filed
    return picked

def get_company_logo(ticker: str, name: str) -> Path | None:
    # Check if local MNDY logo was cached
    if ticker == "MNDY":
        p = ROOT / "Assets & Media" / "Logos" / "MNDY_logo.png"
        if p.exists():
            return p
    
    # Check if we already have it in assets
    local_path = ROOT / "Assets & Media" / "Logos" / f"{ticker}_logo.png"
    if local_path.exists():
        return local_path
        
    # Determine domain
    domain = TICKER_DOMAINS.get(ticker)
    if not domain:
        if "." in name:
            match = re.search(r"([a-zA-Z0-9\-]+\.[a-zA-Z]{2,})", name)
            if match:
                domain = match.group(1).lower()
        if not domain:
            domain = f"{ticker.lower()}.com"
            
    # Try downloading from clearbit
    url = f"https://logo.clearbit.com/{domain}"
    try:
        urllib.request.urlretrieve(url, str(local_path))
        print(f"Downloaded logo for {ticker} from {url}")
        return local_path
    except Exception as e:
        print(f"Failed to download logo for {ticker}: {e}")
        return None

def write_borders(ws, start_row, start_col, end_row, end_col, style="thin", color="808080"):
    side = Side(style=style, color=color)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            border_kwargs = {}
            if r == start_row: border_kwargs["top"] = side
            if r == end_row: border_kwargs["bottom"] = side
            if c == start_col: border_kwargs["left"] = side
            if c == end_col: border_kwargs["right"] = side
            cell.border = Border(**border_kwargs)

def apply_base_styles(ws):
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

def build_dcf_workbook(entity_id: str, method: str = "cash_flow") -> Path:
    node = load_node(entity_id)
    facts, facts_path = load_facts(node)
    ticker = node.get("t") or node["id"]
    
    # Retrieve historical years (last 5 years)
    revenue_series = get_latest_filed_annual(facts, TAGS["revenue"])
    years = sorted(list(revenue_series.keys()))[-5:]
    if len(years) < 2:
        raise ValueError("Not enough annual facts to build DCF")
        
    wb = Workbook()
    wb.remove(wb.active)
    
    # Get logos
    logo_path = get_company_logo(ticker, node.get("n", ""))
    oasis_logo_path = ROOT / "Assets & Media" / "Logos" / "Logo_Dark_BG.png"
    
    # ----------------------------------------------------
    # 1. Cover Sheet
    # ----------------------------------------------------
    ws_cover = wb.create_sheet("Cover")
    apply_base_styles(ws_cover)
    
    # Column Widths
    ws_cover.column_dimensions["A"].width = 10.8
    ws_cover.column_dimensions["B"].width = 13.0
    ws_cover.column_dimensions["C"].width = 59.2
    ws_cover.column_dimensions["D"].width = 10.8
    ws_cover.column_dimensions["J"].width = 13.0
    
    # Put x in A2, XFD2, A13, XFD13, A30, XFD30 to force range (same as user sheet)
    ws_cover.cell(2, 1, "x").font = Font(color=WHITE)
    ws_cover.cell(2, 16384, "x").font = Font(color=WHITE)
    ws_cover.cell(13, 1, "x").font = Font(color=WHITE)
    ws_cover.cell(13, 16384, "x").font = Font(color=WHITE)
    ws_cover.cell(30, 1, "x").font = Font(color=WHITE)
    ws_cover.cell(30, 16384, "x").font = Font(color=WHITE)
    
    # Add logos to Cover
    if logo_path and logo_path.exists():
        try:
            img = OpenpyxlImage(str(logo_path))
            # Set size matching user banner
            img.width, img.height = 364, 105
            ws_cover.add_image(img, "C6")
        except Exception as e:
            print(f"Error adding company logo: {e}")
            
    if oasis_logo_path.exists():
        try:
            img_oasis = OpenpyxlImage(str(oasis_logo_path))
            img_oasis.width, img_oasis.height = 418, 235
            ws_cover.add_image(img_oasis, "G2")
        except Exception as e:
            print(f"Error adding Oasis logo: {e}")
            
    # Sheet Headers
    c13 = ws_cover.cell(13, 3, f"{ticker} DCF")
    c13.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    c13.fill = PatternFill("solid", fgColor=DARK_BG)
    
    c15 = ws_cover.cell(15, 3, "Index")
    c15.font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    c15.fill = PatternFill("solid", fgColor=DARK_BG)
    
    # Index Links
    index_rows = [
        ("DCF - Financials", f"Historical Financial Statements for {ticker}"),
        ("DCF - Assumptions", "Assumptions and inputs"),
        ("DCF - Assumptions Document", "Summary of key assumptions"),
        ("DCF - Valuation", "Valuation assumptions and calculations")
    ]
    for idx, (sh_name, sh_desc) in enumerate(index_rows, 16):
        ws_cover.cell(idx, 3, sh_name).font = Font(name="Aptos Narrow", size=12, color="0000FF", underline="single")
        ws_cover.cell(idx, 7, sh_desc).font = Font(name="Aptos Narrow", size=12, color="666666")
        
    # Findings Summary
    c21 = ws_cover.cell(21, 3, "Findings")
    c21.font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    c21.fill = PatternFill("solid", fgColor=DARK_BG)
    
    findings = [
        ("Enterprise Value ", f"='DCF - Valuation.{ticker}'!P117", USD_MM),
        ("   Cash ", f"='DCF - Valuation.{ticker}'!P118", USD_MM),
        ("   Total Debt ", f"='DCF - Valuation.{ticker}'!P119", USD_MM),
        ("   Equity Value ", f"='DCF - Valuation.{ticker}'!P120", USD_MM),
        ("   Total Share Outstanding ", f"='DCF - Valuation.{ticker}'!P122", NUM),
        ("Implied Share Price", f"='DCF - Valuation.{ticker}'!P123", USD_MM),
        ("Current Share Price", f"='DCF - Valuation.{ticker}'!K14", USD_MM),
        ("Downside", f"='DCF - Valuation.{ticker}'!K15", PCT)
    ]
    for idx, (label, formula, num_fmt) in enumerate(findings, 22):
        ws_cover.cell(idx, 3, label).font = Font(name="Aptos Narrow", size=12, bold=("Price" in label or "Downside" in label))
        cell_val = ws_cover.cell(idx, 10, formula)
        cell_val.font = Font(name="Aptos Narrow", size=12)
        cell_val.number_format = num_fmt
        cell_val.alignment = Alignment(horizontal="right")
        
    write_borders(ws_cover, 27, 3, 29, 10, style="thin")
    
    # ----------------------------------------------------
    # 2. Financials Sheet
    # ----------------------------------------------------
    ws_fin = wb.create_sheet(f"DCF - Financials.{ticker}")
    apply_base_styles(ws_fin)
    ws_fin.column_dimensions["B"].width = 53.0
    ws_fin.column_dimensions["K"].width = 37.6
    ws_fin.column_dimensions["T"].width = 44.6
    for col in ["C", "D", "E", "F", "G", "H", "I", "L", "M", "N", "O", "P", "Q", "R", "U", "V", "W", "X", "Y", "Z"]:
        ws_fin.column_dimensions[col].width = 13.0
        
    # Banner at Row 2
    ws_fin.cell(2, 2, "Figures are in millions (M) except for EPS which is in $/share").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_fin.cell(2, 2).fill = PatternFill("solid", fgColor=DARK_BG)
    for col_idx in range(3, 27):
        ws_fin.cell(2, col_idx).fill = PatternFill("solid", fgColor=DARK_BG)
        
    # Add Oasis logo at W1
    if oasis_logo_path.exists():
        try:
            img_oasis = OpenpyxlImage(str(oasis_logo_path))
            img_oasis.width, img_oasis.height = 209, 118
            ws_fin.add_image(img_oasis, "W1")
        except Exception as e:
            print(f"Error adding Oasis logo: {e}")
            
    # Extract values for columns
    hist_years = years[::-1] # latest year first
    
    def fill_column_data(col_start_idx, year_vals, row_map, scale=1e6):
        for idx, y in enumerate(hist_years):
            col_letter = get_column_letter(col_start_idx + idx)
            ws_fin.cell(5, col_start_idx + idx, f"DEC '{str(y)[2:]}").font = Font(name="Aptos Narrow", size=12)
            for row_num, key in row_map.items():
                val = year_vals.get(key, {}).get(y)
                if val is not None:
                    cell = ws_fin.cell(row_num, col_start_idx + idx, val / scale if scale else val)
                    cell.font = Font(name="Aptos Narrow", size=12, color=DARK_BG)
                    cell.alignment = Alignment(horizontal="right")
                    
    # Three Statement Maps
    is_map = {
        8: "sales", 9: "cogs", 11: "da", 15: "sga", 16: "rd", 18: "ebit", 24: "pretax", 25: "tax", 26: "net_income", 27: "net_income", 38: "shares_diluted"
    }
    cf_map = {
        8: "net_income", 9: "da", 18: "cfo", 20: "capex", 24: "cfi", 29: "cff", 26: "dividends"
    }
    bs_map = {
        9: "cash", 11: "st_investments", 13: "receivables", 20: "current_assets", 21: "ppe_net", 28: "total_assets", 31: "ap", 30: "st_debt", 35: "current_liabilities", 36: "lt_debt", 40: "total_liabilities", 47: "equity", 48: "total_assets"
    }
    
    # Extract historical dictionaries
    hist_data = {}
    for key, tags in TAGS.items():
        hist_data[key] = get_latest_filed_annual(facts, tags)
        
    # Write Labels
    ws_fin.cell(6, 2, "GAAP/IFRS Income Statement").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_fin.cell(6, 11, "GAAP/IFRS Cash Flow ").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_fin.cell(6, 20, "GAAP/IFRS Balance Sheet").font = Font(name="Aptos Narrow", size=12, bold=True)
    
    is_labels = {
        8: "Sales", 9: "Cost of Goods Sold (COGS) incl. D&A", 10: "COGS excluding D&A", 11: "Depreciation & Amortization Expense", 12: "Depreciation",
        13: "Amortization of Deferred Charges", 14: "Gross Income", 15: "SG&A Expense", 16: "Research & Development", 17: "Other SG&A",
        18: "EBIT (Operating Income)", 19: "Nonoperating Income - Net", 24: "Pretax Income", 25: "Income Taxes", 26: "Consolidated Net Income",
        27: "Net Income", 29: "Net Income available to Common", 37: "EPS (diluted)", 38: "Diluted Shares Outstanding"
    }
    for r, lbl in is_labels.items():
        ws_fin.cell(r, 2, lbl).font = Font(name="Aptos Narrow", size=12)
        
    cf_labels = {
        7: "Operating Activities", 8: "Net Income / Starting Line", 9: "Depreciation, Depletion & Amortization", 13: "Changes in Working Capital",
        18: "Net Operating Cash Flow", 19: "Investing Activities", 20: "Capital Expenditures", 24: "Net Investing Cash Flow",
        25: "Financing Activities", 26: "Cash Dividends Paid", 29: "Net Financing Cash Flow", 32: "Net Change in Cash"
    }
    for r, lbl in cf_labels.items():
        ws_fin.cell(r, 11, lbl).font = Font(name="Aptos Narrow", size=12)
        
    bs_labels = {
        8: "Assets", 9: "Cash & Short-Term Investments", 10: "Cash Only", 11: "Total Short Term Investments", 12: "Short-Term Receivables",
        13: "Accounts Receivables, Net", 20: "Total Current Assets", 21: "Net Property, Plant & Equipment", 28: "Total Assets",
        29: "Liabilities & Equity", 30: "Short-Term Debt", 31: "Accounts Payable", 35: "Total Current Liabilities", 36: "Long-Term Debt",
        40: "Total Liabilities", 41: "Shareholders Equity", 47: "Total Shareholders Equity", 48: "Total Liabilities & Equity"
    }
    for r, lbl in bs_labels.items():
        ws_fin.cell(r, 20, lbl).font = Font(name="Aptos Narrow", size=12)
        
    # Write Values
    fill_column_data(4, hist_data, is_map)
    fill_column_data(13, hist_data, cf_map)
    fill_column_data(21, hist_data, bs_map)
    
    # ----------------------------------------------------
    # 3. Assumptions Sheet
    # ----------------------------------------------------
    ws_ass = wb.create_sheet(f"DCF - Assumptions.{ticker}")
    apply_base_styles(ws_ass)
    ws_ass.column_dimensions["B"].width = 32.0
    ws_ass.column_dimensions["C"].width = 24.0
    for col in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
        ws_ass.column_dimensions[col].width = 13.0
        
    # Add logos to Assumptions
    if logo_path and logo_path.exists():
        try:
            img = OpenpyxlImage(str(logo_path))
            img.width, img.height = 364, 105
            ws_ass.add_image(img, "C6")
        except Exception as e:
            print(f"Error adding company logo: {e}")
            
    if oasis_logo_path.exists():
        try:
            img_oasis = OpenpyxlImage(str(oasis_logo_path))
            img_oasis.width, img_oasis.height = 418, 235
            ws_ass.add_image(img_oasis, "G2")
        except Exception as e:
            print(f"Error adding Oasis logo: {e}")
            
    ws_ass.cell(5, 3, f"Assumption Models - {ticker}  ").font = Font(name="Aptos Narrow", size=12, bold=True)
    
    # Switches
    ws_ass.cell(7, 3, "Switches").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_ass.cell(7, 3).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_ass.cell(8, 3, "Assumptions").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_ass.cell(8, 6, "Metric").font = Font(name="Aptos Narrow", size=12, bold=True)
    
    switches = [
        ("Revenue/Sales", 9),
        ("EBITDA", 10),
        ("EBIT", 11),
        ("Non-GAAP D&A", 12),
        ("CapEx", 13),
        ("Change In Networking Capital", 14),
        ("Tax", 15)
    ]
    for label, r in switches:
        ws_ass.cell(r, 3, label).font = Font(name="Aptos Narrow", size=12)
        cell_switch = ws_ass.cell(r, 6, 2)
        cell_switch.font = Font(name="Calibri", size=12)
        cell_switch.fill = PatternFill("solid", fgColor=YELLOW_INPUT)
        cell_switch.alignment = Alignment(horizontal="center")
        
    write_borders(ws_ass, 9, 6, 15, 6, style="thin")
    
    # Scenario offsets
    ws_ass.cell(17, 3, "Conservative").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_ass.cell(17, 3).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_ass.cell(17, 8, "Base").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_ass.cell(17, 8).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_ass.cell(17, 13, "Optimistic").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_ass.cell(17, 13).fill = PatternFill("solid", fgColor=DARK_BG)
    
    offset_rows = [
        ("Revenue", "'25-'29", -0.05, "Revenue", "'25-'29", -0.045, "Revenue", "'25-'29", 0.02, 19),
        ("EBITDA", "'25-'29", -0.08, "GAAP EBIT", "'28-'29", -0.08, "EBITDA", "'25-'29", 0.07, 20),
        ("GAAP EBIT", "'25-'26", 0.05, "Non-GAAP EBIT", "'28-'29", -0.10, "GAAP EBIT", "'25-'26", -0.05, 21),
        ("GAAP EBIT", "'27-'29", -0.025, "", "", "", "GAAP EBIT", "'27-'29", 0.10, 22),
        ("Non-GAAP EBIT", "'25-'29", -0.07, "", "", "", "Non-GAAP EBIT", "'25-'29", 0.10, 23),
        ("Non-GAAP D&A", "'25-'27", 0.02, "", "", "", "Change In Networking Capital", "'25-'29", -0.07, 24),
        ("Non-GAAP D&A", "'28-'29", 0.03, "", "", "", "CapEx", "'25-'29", -0.07, 25),
    ]
    for r_data in offset_rows:
        r = r_data[9]
        ws_ass.cell(r, 3, r_data[0]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 5, r_data[1]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 6, r_data[2]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 6).number_format = PCT
        ws_ass.cell(r, 8, r_data[3]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 10, r_data[4]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 11, r_data[5]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 11).number_format = PCT
        ws_ass.cell(r, 13, r_data[6]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 15, r_data[7]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 16, r_data[8]).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(r, 16).number_format = PCT
        
    # Write Operating Build
    ws_ass.cell(30, 2, "Operating Build YoY").font = Font(name="Aptos Narrow", size=12, bold=True)
    
    # Forecast years headers
    ws_ass.cell(32, 2, "Revenue").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_ass.cell(32, 2).fill = PatternFill("solid", fgColor=DARK_BG)
    for col_idx, y in enumerate(years):
        ws_ass.cell(32, 8 + col_idx, y).font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
        ws_ass.cell(32, 8 + col_idx).fill = PatternFill("solid", fgColor=DARK_BG)
        
    for idx in range(5):
        y_f = years[-1] + 1 + idx
        ws_ass.cell(32, 13 + idx, y_f).font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
        ws_ass.cell(32, 13 + idx).fill = PatternFill("solid", fgColor=DARK_BG)
        
    # Put values & formulas for Operating Build
    rev_hist = [hist_data["revenue"].get(y, 0) / 1e6 for y in years]
    rev_cagr = (rev_hist[-1] / rev_hist[0]) ** (1 / (len(years) - 1)) - 1 if rev_hist[0] else 0.05
    rev_cagr = min(0.35, max(0.05, rev_cagr))
    
    # Revenue rows (Row 33-37)
    ws_ass.cell(33, 2, node.get("n", "Revenue")).font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        cell = ws_ass.cell(33, 8 + idx, rev_hist[idx])
        cell.font = Font(name="Aptos Narrow", size=12, color=BLUE_TEXT)
        cell.number_format = USD_MM
        
    for idx in range(5):
        col_letter = get_column_letter(13 + idx)
        prev_col = get_column_letter(12 + idx)
        ws_ass.cell(33, 13 + idx, f"={prev_col}33*(1+{col_letter}34)").font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(33, 13 + idx).number_format = USD_MM
        
    # % growth row (Row 34)
    ws_ass.cell(34, 2, "% growth").font = Font(name="Aptos Narrow", size=12)
    ws_ass.cell(34, 8, "--").font = Font(name="Aptos Narrow", size=12)
    for idx in range(1, len(years)):
        col_letter = get_column_letter(8 + idx)
        prev_col = get_column_letter(7 + idx)
        ws_ass.cell(34, 8 + idx, f"={col_letter}33/{prev_col}33-1").font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(34, 8 + idx).number_format = PCT
        
    for idx in range(5):
        col_letter = get_column_letter(13 + idx)
        ws_ass.cell(34, 13 + idx, f"=OFFSET({col_letter}34,$F$9,0)").font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(34, 13 + idx).number_format = PCT
        
    # Cases
    for idx in range(5):
        col_letter = get_column_letter(13 + idx)
        val_street = rev_cagr * (0.9 ** idx)
        ws_ass.cell(36, 13 + idx, val_street).font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(36, 13 + idx).number_format = PCT
        ws_ass.cell(35, 13 + idx, f"={col_letter}36*(1+$F$19)").font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(35, 13 + idx).number_format = PCT
        ws_ass.cell(37, 13 + idx, f"={col_letter}36*(1+$P$19)").font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(37, 13 + idx).number_format = PCT
        
    operating_metrics = [
        ("EBITDA", "da", 39, 40, 41, 42, 43, 44, "$F$10", "$P$20", 0.15),
        ("GAAP EBIT", "ebit", 46, 47, 48, 49, 50, 51, "$F$21", "$P$21", 0.10),
        ("Non-GAAP EBIT", "ebit", 53, 54, 55, 56, 57, 58, "$F$23", "$P$23", 0.12),
        ("D&A", "da", 60, 61, 62, 63, 64, 65, "$F$24", "$P$24", 0.05),
        ("CapEx", "capex", 67, 68, 69, 70, 71, 72, "$F$27", "$P$25", 0.06),
        ("Change In Networking Capital", "cfo", 74, 75, 76, 77, 78, 79, "$F$26", "$P$24", 0.02)
    ]
    
    for lbl, fact_key, hr, vr, pr, cr, sr, or_row, cons_mult, opt_mult, default_rate in operating_metrics:
        ws_ass.cell(hr, 2, lbl).font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
        ws_ass.cell(hr, 2).fill = PatternFill("solid", fgColor=DARK_BG)
        for col_idx in range(8, 18):
            ws_ass.cell(hr, col_idx).fill = PatternFill("solid", fgColor=DARK_BG)
            
        ws_ass.cell(vr, 2, lbl).font = Font(name="Aptos Narrow", size=12)
        for idx, y in enumerate(years):
            val = hist_data[fact_key].get(y, 0) / 1e6
            cell = ws_ass.cell(vr, 8 + idx, val)
            cell.font = Font(name="Aptos Narrow", size=12, color=BLUE_TEXT)
            cell.number_format = USD_MM
            
        for idx in range(5):
            col_letter = get_column_letter(13 + idx)
            prev_col = get_column_letter(12 + idx)
            ws_ass.cell(vr, 13 + idx, f"={prev_col}33*(1+{col_letter}{pr})").font = Font(name="Aptos Narrow", size=12)
            ws_ass.cell(vr, 13 + idx).number_format = USD_MM
            
        # % growth
        ws_ass.cell(pr, 2, "% growth").font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(pr, 8, "--").font = Font(name="Aptos Narrow", size=12)
        for idx in range(1, len(years)):
            col_letter = get_column_letter(8 + idx)
            prev_col = get_column_letter(7 + idx)
            ws_ass.cell(pr, 8 + idx, f"={col_letter}{vr}/{prev_col}{vr}-1").font = Font(name="Aptos Narrow", size=12)
            ws_ass.cell(pr, 8 + idx).number_format = PCT
            
        for idx in range(5):
            col_letter = get_column_letter(13 + idx)
            ws_ass.cell(pr, 13 + idx, f"=OFFSET({col_letter}{pr},{cons_mult},0)").font = Font(name="Aptos Narrow", size=12)
            ws_ass.cell(pr, 13 + idx).number_format = PCT
            
        # Cases
        for idx in range(5):
            col_letter = get_column_letter(13 + idx)
            val_street = default_rate * (0.95 ** idx)
            ws_ass.cell(sr, 13 + idx, val_street).font = Font(name="Aptos Narrow", size=12)
            ws_ass.cell(sr, 13 + idx).number_format = PCT
            ws_ass.cell(cr, 13 + idx, f"={col_letter}{sr}*(1+{cons_mult})").font = Font(name="Aptos Narrow", size=12)
            ws_ass.cell(cr, 13 + idx).number_format = PCT
            ws_ass.cell(or_row, 13 + idx, f"={col_letter}{sr}*(1+{opt_mult})").font = Font(name="Aptos Narrow", size=12)
            ws_ass.cell(or_row, 13 + idx).number_format = PCT
            
    # Write Tax rates
    ws_ass.cell(81, 2, "Tax Rate").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_ass.cell(81, 2).fill = PatternFill("solid", fgColor=DARK_BG)
    for col_idx in range(8, 18):
        ws_ass.cell(81, col_idx).fill = PatternFill("solid", fgColor=DARK_BG)
        
    for idx, y in enumerate(years):
        ws_ass.cell(81, 8 + idx, f"{y}A").font = Font(name="Aptos Narrow", size=12)
        
    ws_ass.cell(82, 2, "Tax").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        val = hist_data["tax"].get(y, 0) / 1e6
        ws_ass.cell(82, 8 + idx, val).font = Font(name="Aptos Narrow", size=12, color=BLUE_TEXT)
        
    ws_ass.cell(83, 2, "Tax Rate").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(13 + idx)
        ws_ass.cell(83, 13 + idx, f"=OFFSET({col_letter}83,$F$15,0)").font = Font(name="Aptos Narrow", size=12)
        ws_ass.cell(83, 13 + idx).number_format = PCT
        
    ws_ass.cell(84, 2, "Tax Rate for GAAP EBIT").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        ws_ass.cell(84, 13 + idx, 0.05).font = Font(name="Aptos Narrow", size=12)
        
    ws_ass.cell(85, 2, "Tax Rate for Non-GAAP EBIT").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        ws_ass.cell(85, 13 + idx, 0.05).font = Font(name="Aptos Narrow", size=12)
        
    # ----------------------------------------------------
    # 4. Assumptions Document Sheet
    # ----------------------------------------------------
    ws_doc = wb.create_sheet(f"DCF - Assumptions Doc.{ticker}")
    apply_base_styles(ws_doc)
    ws_doc.column_dimensions["C"].width = 24.0
    ws_doc.column_dimensions["D"].width = 13.0
    ws_doc.column_dimensions["E"].width = 18.0
    ws_doc.column_dimensions["F"].width = 72.5
    
    ws_doc.cell(2, 1, "x").font = Font(color=WHITE)
    ws_doc.cell(2, 16251, "x").font = Font(color=WHITE)
    ws_doc.cell(13, 1, "x").font = Font(color=WHITE)
    ws_doc.cell(13, 16251, "x").font = Font(color=WHITE)
    ws_doc.cell(24, 1, "x").font = Font(color=WHITE)
    ws_doc.cell(24, 16251, "x").font = Font(color=WHITE)
    
    if logo_path and logo_path.exists():
        try:
            img = OpenpyxlImage(str(logo_path))
            img.width, img.height = 364, 105
            ws_doc.add_image(img, "C6")
        except Exception as e:
            print(f"Error adding company logo: {e}")
            
    if oasis_logo_path.exists():
        try:
            img_oasis = OpenpyxlImage(str(oasis_logo_path))
            img_oasis.width, img_oasis.height = 418, 235
            ws_doc.add_image(img_oasis, "G2")
        except Exception as e:
            print(f"Error adding Oasis logo: {e}")
            
    ws_doc.cell(13, 3, "Summary of Key Assumptions ").font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws_doc.cell(13, 3).fill = PatternFill("solid", fgColor=DARK_BG)
    
    c15 = ws_doc.cell(15, 3, "Discounted Cashflow Valuation (DCF) Assumptions")
    c15.font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    c15.fill = PatternFill("solid", fgColor=DARK_BG)
    
    ws_doc.cell(16, 3, "Metric").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_doc.cell(16, 4, "Weight").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_doc.cell(16, 5, "Scenario").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_doc.cell(16, 6, "Comment").font = Font(name="Aptos Narrow", size=12, bold=True)
    
    doc_rows = [
        ("Revenue", 3, "Optimistic", "Reflects strong revenue growth"),
        ("EBIT", 1, "Conservative", "Accounts for operating losses despite using Non-GAAP figures"),
        ("D&A", 2, "Base", "Assumes average depreciation/amortization based on company metrics"),
        ("CapEX", 1, "Conservative", "Reflects cautious investment in long-term assets amid growth"),
        ("Change in NWC", 1, "Conservative", "Accounts conservatively for operational fluctuations, including interest income effects"),
        ("WACC", 1, "Conservative", "Aligned with analyst recommendations"),
        ("TGR", 2, "Base", "Uses economy-wide growth as the company’s GAAP EBIT loss limits high long-term growth")
    ]
    for idx, (metric, weight, scenario, comment) in enumerate(doc_rows, 17):
        ws_doc.cell(idx, 3, metric).font = Font(name="Aptos Narrow", size=12)
        cell_w = ws_doc.cell(idx, 4, weight)
        cell_w.font = Font(name="Aptos Narrow", size=12)
        cell_w.alignment = Alignment(horizontal="center")
        ws_doc.cell(idx, 5, scenario).font = Font(name="Aptos Narrow", size=12)
        ws_doc.cell(idx, 6, comment).font = Font(name="Aptos Narrow", size=12)
        
    write_borders(ws_doc, 17, 3, 23, 6, style="thin")
    
    # ----------------------------------------------------
    # 5. Valuation Sheet
    # ----------------------------------------------------
    ws_val = wb.create_sheet(f"DCF - Valuation.{ticker}")
    apply_base_styles(ws_val)
    ws_val.column_dimensions["C"].width = 47.0
    ws_val.column_dimensions["D"].width = 12.2
    ws_val.column_dimensions["H"].width = 11.0
    ws_val.column_dimensions["K"].width = 13.0
    ws_val.column_dimensions["M"].width = 11.7
    ws_val.column_dimensions["P"].width = 13.2
    for col in ["E", "F", "G", "I", "J", "L", "N", "O", "Q", "R", "S", "T", "U", "V", "W"]:
        ws_val.column_dimensions[col].width = 13.0
        
    ws_val.cell(2, 1, "x").font = Font(color=WHITE)
    ws_val.cell(2, 18, "x").font = Font(color=WHITE)
    ws_val.cell(35, 1, "x").font = Font(color=WHITE)
    ws_val.cell(35, 17, "x").font = Font(color=WHITE)
    ws_val.cell(72, 1, "x").font = Font(color=WHITE)
    ws_val.cell(72, 17, "x").font = Font(color=WHITE)
    ws_val.cell(111, 1, "x").font = Font(color=WHITE)
    ws_val.cell(111, 17, "x").font = Font(color=WHITE)
    
    if oasis_logo_path.exists():
        try:
            img_oasis = OpenpyxlImage(str(oasis_logo_path))
            img_oasis.width, img_oasis.height = 418, 235
            ws_val.add_image(img_oasis, "G2")
        except Exception as e:
            print(f"Error adding Oasis logo: {e}")
            
    ws_val.cell(5, 3, f"Discounted Cashflow Valuation (DCF) Model - {ticker}  ").font = Font(name="Aptos Narrow", size=12, bold=True)
    
    ws_val.cell(7, 3, "WACC").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(7, 3).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_val.cell(7, 4, "($US millions)").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(7, 4).fill = PatternFill("solid", fgColor=DARK_BG)
    
    price_dict = json.load(DATA.joinpath("prices.json").open()) if DATA.joinpath("prices.json").exists() else {}
    company_price = price_dict.get(ticker, {}).get("price", 185.03)
    shares_latest = hist_data["shares"].get(years[-1], 53.3e6)
    mcap = company_price * shares_latest / 1e6
    
    wacc_data = [
        ("Market Capitalization", mcap, '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'),
        ("Debt", hist_data["debt"].get(years[-1], 77e6) / 1e6, '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'),
        ("Risk Free Rate(5yr treasury)", 0.04059, "0.00%"),
        ("Equity Risk Premium (Damodaran)", 0.0377, "0.00%"),
        ("Beta", 1.79, "0.00"),
        ("Cost of Equity", "=D11+D13*(D12)", "0.00%"),
        ("Weight of Equity", "=D9/(D9+D10)", "0.00%"),
        ("Cost of Debt", 0.088, "0.00%"),
        ("Weight of Debt", "=D10/(D9+D10)", "0.00%"),
        ("Cost of Preferred Stock", 0.0, "0.00%"),
        ("Weight of Preferred Stock", 0.0, "0.00%"),
        ("Tax Rate", 0.05, "0.00%"),
        ("WACC", "=(D15*D14)+(D18*D17)*(1-D20)", "0.00%")
    ]
    for idx, (lbl, val, num_fmt) in enumerate(wacc_data, 9):
        ws_val.cell(idx, 3, lbl).font = Font(name="Aptos Narrow", size=12)
        cell_val = ws_val.cell(idx, 4, val)
        cell_val.font = Font(name="Aptos Narrow", size=12)
        cell_val.number_format = num_fmt
        if isinstance(val, (int, float)):
            cell_val.alignment = Alignment(horizontal="right")
            
    ws_val.cell(23, 3, "Terminal Growth Rate").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_val.cell(23, 4, 0.025).font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(23, 4).number_format = "0.00%"
    
    ws_val.cell(7, 8, "Dashboard").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(7, 8).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_val.cell(9, 8, "Ticker").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(9, 11, ticker).font = Font(name="Calibri", size=12)
    ws_val.cell(9, 11).alignment = Alignment(horizontal="center")
    
    ws_val.cell(10, 8, "Date").font = Font(name="Aptos Narrow", size=12)
    cell_date = ws_val.cell(10, 11, datetime.now())
    cell_date.font = Font(name="Calibri", size=12)
    cell_date.number_format = "mm-dd-yy"
    cell_date.alignment = Alignment(horizontal="center")
    
    ws_val.cell(11, 8, "Year End").font = Font(name="Aptos Narrow", size=12)
    cell_ye = ws_val.cell(11, 11, datetime(datetime.now().year, 12, 31))
    cell_ye.font = Font(name="Calibri", size=12)
    cell_ye.number_format = "mm-dd-yy"
    cell_ye.alignment = Alignment(horizontal="center")
    
    ws_val.cell(13, 8, "Implied Price Per Share").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(13, 11, "=P123").font = Font(name="Calibri", size=12)
    ws_val.cell(13, 11).number_format = USD_MM
    
    ws_val.cell(14, 8, "Current Share Price").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(14, 11, company_price).font = Font(name="Calibri", size=12)
    ws_val.cell(14, 11).number_format = USD_MM
    
    ws_val.cell(15, 8, "Implied Upside/(Downside)").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(15, 11, "=K13/K14-1").font = Font(name="Calibri", size=12)
    ws_val.cell(15, 11).number_format = PCT
    
    write_borders(ws_val, 9, 11, 15, 11, style="thin")
    
    ws_val.cell(7, 13, "Switches").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(7, 13).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_val.cell(8, 13, "Assumptions").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_val.cell(9, 13, "Revenue").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(9, 16, f"='DCF - Assumptions.{ticker}'!F9").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(10, 13, "EBIT").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(10, 16, f"='DCF - Assumptions.{ticker}'!F11").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(11, 13, "D&A").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(11, 16, f"='DCF - Assumptions.{ticker}'!F12").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(12, 13, "CapEX").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(12, 16, f"='DCF - Assumptions.{ticker}'!F13").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(13, 13, "Change in NWC").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(13, 16, f"='DCF - Assumptions.{ticker}'!F14").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(14, 13, "WACC").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(14, 16, f"='DCF - Assumptions Doc.{ticker}'!D22").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(15, 13, "TGR").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(15, 16, f"='DCF - Assumptions Doc.{ticker}'!D23").font = Font(name="Aptos Narrow", size=12)
    
    ws_val.cell(17, 13, "Valuation").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_val.cell(18, 13, "WACC").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(18, 16, "=CHOOSE(P14,F27,K27,P27)").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(18, 16).number_format = PCT
    ws_val.cell(19, 13, "TGR").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(19, 16, "=CHOOSE(P15,F28,K28,P28)").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(19, 16).number_format = PCT
    
    write_borders(ws_val, 9, 16, 15, 16, style="thin")
    write_borders(ws_val, 18, 16, 19, 16, style="thin")
    
    ws_val.cell(25, 3, "Conservative").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(25, 3).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_val.cell(25, 8, "Base").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(25, 8).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_val.cell(25, 13, "Optimistic").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(25, 13).fill = PatternFill("solid", fgColor=DARK_BG)
    
    ws_val.cell(26, 3, "Assumptions").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_val.cell(26, 6, "Metric").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_val.cell(26, 8, "Assumptions").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_val.cell(26, 11, "Metric").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_val.cell(26, 13, "Assumptions").font = Font(name="Aptos Narrow", size=12, bold=True)
    ws_val.cell(26, 16, "Metric").font = Font(name="Aptos Narrow", size=12, bold=True)
    
    # Cons
    ws_val.cell(27, 3, "WACC").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(27, 6, "=D21+0.005").font = Font(name="Calibri", size=12)
    ws_val.cell(27, 6).number_format = PCT
    ws_val.cell(28, 3, "TGR").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(28, 6, 0.02).font = Font(name="Calibri", size=12)
    ws_val.cell(28, 6).number_format = PCT
    # Base
    ws_val.cell(27, 8, "WACC").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(27, 11, "=D21").font = Font(name="Calibri", size=12)
    ws_val.cell(27, 11).number_format = PCT
    ws_val.cell(28, 8, "TGR").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(28, 11, 0.025).font = Font(name="Calibri", size=12)
    ws_val.cell(28, 11).number_format = PCT
    # Opt
    ws_val.cell(27, 13, "WACC").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(27, 16, "=D21-0.005").font = Font(name="Calibri", size=12)
    ws_val.cell(27, 16).number_format = PCT
    ws_val.cell(28, 13, "TGR").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(28, 16, 0.03).font = Font(name="Calibri", size=12)
    ws_val.cell(28, 16).number_format = PCT
    
    write_borders(ws_val, 27, 6, 28, 6, style="thin")
    write_borders(ws_val, 27, 11, 28, 11, style="thin")
    write_borders(ws_val, 27, 16, 28, 16, style="thin")
    
    ws_val.cell(35, 2, "Historicals Financials Items").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(35, 2).fill = PatternFill("solid", fgColor=DARK_BG)
    ws_val.cell(35, 16, "($US millions)").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(35, 16).fill = PatternFill("solid", fgColor=DARK_BG)
    
    ws_val.cell(37, 2, "Income Statement Items").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(37, 2).fill = PatternFill("solid", fgColor=DARK_BG)
    for idx, y in enumerate(years):
        ws_val.cell(37, 8 + idx, f"{y}A").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
        ws_val.cell(37, 8 + idx).fill = PatternFill("solid", fgColor=DARK_BG)
        
    # Revenue (Row 38)
    ws_val.cell(38, 2, "Revenue").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(38, 8 + idx, f"='DCF - Assumptions.{ticker}'!{col_letter}33").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(38, 8 + idx).number_format = USD_MM
        
    ws_val.cell(39, 2, "% growth").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(39, 8, "--").font = Font(name="Aptos Narrow", size=12)
    for idx in range(1, len(years)):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(39, 8 + idx, f"=OFFSET({col_letter}39,-1,0)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(39, 8 + idx).number_format = PCT
        
    # EBITDA
    ws_val.cell(41, 2, "EBITDA").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(41, 8 + idx, f"='DCF - Assumptions.{ticker}'!{col_letter}40").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(41, 8 + idx).number_format = USD_MM
        
    ws_val.cell(42, 2, "D&A").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(42, 8 + idx, f"='DCF - Assumptions.{ticker}'!{col_letter}61").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(42, 8 + idx).number_format = USD_MM
        
    ws_val.cell(43, 2, "EBIT").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(43, 8 + idx, f"=H41-H42").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(43, 8 + idx).number_format = USD_MM
        
    ws_val.cell(44, 2, "% of sales").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(44, 8 + idx, f"={col_letter}43/{col_letter}38").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(44, 8 + idx).number_format = PCT
        
    ws_val.cell(45, 2, "% of sales adjusted").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(45, 8 + idx, f"={col_letter}44*(-1)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(45, 8 + idx).number_format = PCT
        
    # Cash Flow Items
    ws_val.cell(60, 2, "Cash Flow Items ").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(60, 2).fill = PatternFill("solid", fgColor=DARK_BG)
    for col_idx in range(8, 18):
        ws_val.cell(60, col_idx).fill = PatternFill("solid", fgColor=DARK_BG)
        
    ws_val.cell(61, 2, "D&A").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(61, 8 + idx, f"=OFFSET({col_letter}61,-19,0)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(61, 8 + idx).number_format = USD_MM
        
    ws_val.cell(62, 2, "% of sales").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(62, 8 + idx, f"={col_letter}61/{col_letter}38").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(62, 8 + idx).number_format = PCT
        
    ws_val.cell(64, 2, "CapEx").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(64, 8 + idx, f"='DCF - Assumptions.{ticker}'!{col_letter}68").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(64, 8 + idx).number_format = USD_MM
        
    ws_val.cell(65, 2, "% of sales").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(65, 8 + idx, f"={col_letter}64/{col_letter}38").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(65, 8 + idx).number_format = PCT
        
    ws_val.cell(66, 2, "% of sales adjusted").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(66, 8 + idx, f"={col_letter}65*(-1)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(66, 8 + idx).number_format = PCT
        
    ws_val.cell(68, 2, "Change in NWC").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(68, 8 + idx, f"='DCF - Assumptions.{ticker}'!{col_letter}75").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(68, 8 + idx).number_format = USD_MM
        
    ws_val.cell(69, 2, "% of sales").font = Font(name="Aptos Narrow", size=12)
    for idx, y in enumerate(years):
        col_letter = get_column_letter(8 + idx)
        ws_val.cell(69, 8 + idx, f"={col_letter}68/{col_letter}38").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(69, 8 + idx).number_format = PCT
        
    ws_val.cell(70, 2, "% of change in sales").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(70, 8, "--").font = Font(name="Aptos Narrow", size=12)
    for idx in range(1, len(years)):
        col_letter = get_column_letter(8 + idx)
        prev_col = get_column_letter(7 + idx)
        ws_val.cell(70, 8 + idx, f"={col_letter}68/({col_letter}38-{prev_col}38)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(70, 8 + idx).number_format = PCT
        
    # DCF forecast header
    ws_val.cell(74, 2, "DCF").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
    ws_val.cell(74, 2).fill = PatternFill("solid", fgColor=DARK_BG)
    for col_idx in range(8, 18):
        ws_val.cell(74, col_idx).fill = PatternFill("solid", fgColor=DARK_BG)
        
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(74, 12 + idx, f"{years[-1] + 1 + idx}E").font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
        ws_val.cell(74, 12 + idx).fill = PatternFill("solid", fgColor=DARK_BG)
        
    # Revenue (Row 75)
    ws_val.cell(75, 2, "Revenue").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(75, 12 + idx, f"=OFFSET({col_letter}75,$P$9,0)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(75, 12 + idx).number_format = USD_MM
        
    ws_val.cell(76, 2, "Conservative Case").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
        ws_val.cell(76, 12 + idx, f"={prev_col}38*(1+'DCF - Assumptions.{ticker}'!{col_letter}35)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(76, 12 + idx).number_format = USD_MM
        
    ws_val.cell(77, 2, "Street Case").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
        ws_val.cell(77, 12 + idx, f"={prev_col}38*(1+'DCF - Assumptions.{ticker}'!{col_letter}36)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(77, 12 + idx).number_format = USD_MM
        
    ws_val.cell(78, 2, "Optimistic Case").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
        ws_val.cell(78, 12 + idx, f"={prev_col}38*(1+'DCF - Assumptions.{ticker}'!{col_letter}37)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(78, 12 + idx).number_format = USD_MM
        
    # Non-GAAP EBIT (Row 80)
    ws_val.cell(80, 2, "Non-GAAP EBIT").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(80, 12 + idx, f"=OFFSET({col_letter}80,$P$10,0)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(80, 12 + idx).number_format = USD_MM
        
    ws_val.cell(81, 2, "Conservative Case").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
        ws_val.cell(81, 12 + idx, f"={prev_col}80*(1+'DCF - Assumptions.{ticker}'!{col_letter}56)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(81, 12 + idx).number_format = USD_MM
        
    ws_val.cell(82, 2, "Street Case").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
        ws_val.cell(82, 12 + idx, f"={prev_col}80*(1+'DCF - Assumptions.{ticker}'!{col_letter}57)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(82, 12 + idx).number_format = USD_MM
        
    ws_val.cell(83, 2, "Optimistic Case").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
        ws_val.cell(83, 12 + idx, f"={prev_col}80*(1+'DCF - Assumptions.{ticker}'!{col_letter}58)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(83, 12 + idx).number_format = USD_MM
        
    ws_val.cell(87, 2, "Taxes").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(87, 12 + idx, f"=-{col_letter}80*{col_letter}85").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(87, 12 + idx).number_format = USD_MM
        
    ws_val.cell(88, 2, "NOPAT").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(88, 12 + idx, f"={col_letter}80+{col_letter}87").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(88, 12 + idx).number_format = USD_MM
        
    # D&A (Row 90)
    ws_val.cell(90, 2, "D&A").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(90, 12 + idx, f"=OFFSET({col_letter}90,$P$11,0)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(90, 12 + idx).number_format = USD_MM
        
    for c_idx, name in [(91, "Conservative Case"), (92, "Street Case"), (93, "Optimistic Case")]:
        ws_val.cell(c_idx, 2, name).font = Font(name="Aptos Narrow", size=12)
        for idx in range(5):
            col_letter = get_column_letter(12 + idx)
            ass_row = 63 if name == "Conservative Case" else (64 if name == "Street Case" else 65)
            prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
            ws_val.cell(c_idx, 12 + idx, f"={prev_col}61*(1+'DCF - Assumptions.{ticker}'!{col_letter}{ass_row})").font = Font(name="Aptos Narrow", size=12)
            ws_val.cell(c_idx, 12 + idx).number_format = USD_MM
            
    # Capex (Row 96)
    ws_val.cell(96, 2, "Capex").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(96, 12 + idx, f"=OFFSET({col_letter}96,$P$12,0)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(96, 12 + idx).number_format = USD_MM
        
    for c_idx, name in [(97, "Conservative Case"), (98, "Street Case"), (99, "Optimistic Case")]:
        ws_val.cell(c_idx, 2, name).font = Font(name="Aptos Narrow", size=12)
        for idx in range(5):
            col_letter = get_column_letter(12 + idx)
            ass_row = 70 if name == "Conservative Case" else (71 if name == "Street Case" else 72)
            prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
            ws_val.cell(c_idx, 12 + idx, f"={prev_col}64*(1+'DCF - Assumptions.{ticker}'!{col_letter}{ass_row})").font = Font(name="Aptos Narrow", size=12)
            ws_val.cell(c_idx, 12 + idx).number_format = USD_MM
            
    # Change in NWC (Row 102)
    ws_val.cell(102, 2, "Change in NWC").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(102, 12 + idx, f"=OFFSET({col_letter}102,$P$13,0)").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(102, 12 + idx).number_format = USD_MM
        
    for c_idx, name in [(103, "Conservative Case"), (104, "Street Case"), (105, "Optimistic Case")]:
        ws_val.cell(c_idx, 2, name).font = Font(name="Aptos Narrow", size=12)
        for idx in range(5):
            col_letter = get_column_letter(12 + idx)
            ass_row = 77 if name == "Conservative Case" else (78 if name == "Street Case" else 79)
            prev_col = get_column_letter(11 + idx) if idx > 0 else "K"
            ws_val.cell(c_idx, 12 + idx, f"={prev_col}68*(1+'DCF - Assumptions.{ticker}'!{col_letter}{ass_row})").font = Font(name="Aptos Narrow", size=12)
            ws_val.cell(c_idx, 12 + idx).number_format = USD_MM
            
    # Unlevered FCF (Row 108)
    ws_val.cell(108, 2, "Unlevered FCF").font = Font(name="Aptos Narrow", size=12)
    for idx in range(5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(108, 12 + idx, f"={col_letter}88+{col_letter}90-{col_letter}96-{col_letter}102").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(108, 12 + idx).number_format = USD_MM
        
    # PV of FCF (Row 109)
    ws_val.cell(109, 2, "Present Value of FCF").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(109, 12, "=(L108*L112)/(1+P18)^L113").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(109, 12).number_format = USD_MM
    for idx in range(1, 5):
        col_letter = get_column_letter(12 + idx)
        ws_val.cell(109, 12 + idx, f"={col_letter}108/(1+$P$18)^{col_letter}113").font = Font(name="Aptos Narrow", size=12)
        ws_val.cell(109, 12 + idx).number_format = USD_MM
        
    ws_val.cell(112, 2, "Period").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(112, 12, "=YEARFRAC(K10,K11)").font = Font(name="Aptos Narrow", size=12)
    
    ws_val.cell(113, 2, "Discount Period").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(113, 12, "=L112/2").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(113, 13, "=L112+0.5").font = Font(name="Aptos Narrow", size=12)
    for idx in range(2, 5):
        col_letter = get_column_letter(12 + idx)
        prev_col = get_column_letter(11 + idx)
        ws_val.cell(113, 12 + idx, f"={prev_col}113+1").font = Font(name="Aptos Narrow", size=12)
        
    ws_val.cell(115, 2, "Terminal Value").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(115, 16, "=(P108*(1+P19))/P18-P19").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(115, 16).number_format = USD_MM
    
    ws_val.cell(116, 2, "Present Value of Terminal Value").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(116, 16, "=P115/(1+P18)^P113").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(116, 16).number_format = USD_MM
    
    ws_val.cell(117, 2, "Enterprise Value").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(117, 16, "=SUM(L109:P109,P116)").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(117, 16).number_format = USD_MM
    
    ws_val.cell(118, 2, "(+) Cash").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(118, 16, hist_data["cash"].get(years[-1], 1461.6e6) / 1e6).font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(118, 16).number_format = USD_MM
    
    ws_val.cell(119, 2, "(-) Debt").font = Font(name="Aptos Narrow", size=12)
    debt_val = hist_data["debt"].get(years[-1], 77e6) / 1e6
    ws_val.cell(119, 16, debt_val).font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(119, 16).number_format = USD_MM
    
    ws_val.cell(120, 2, "Equity Value").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(120, 16, "=P117+P118-P119").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(120, 16).number_format = USD_MM
    
    ws_val.cell(122, 2, "Diluted Shares").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(122, 16, shares_latest / 1e6).font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(122, 16).number_format = NUM
    
    ws_val.cell(123, 2, "Implied Stock Price").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(123, 16, "=P120/P122").font = Font(name="Aptos Narrow", size=12)
    ws_val.cell(123, 16).number_format = USD_MM
    
    write_borders(ws_val, 75, 12, 75, 16, style="thin")
    write_borders(ws_val, 80, 12, 80, 16, style="thin")
    write_borders(ws_val, 90, 12, 90, 16, style="thin")
    write_borders(ws_val, 96, 12, 96, 16, style="thin")
    write_borders(ws_val, 102, 12, 102, 16, style="thin")
    write_borders(ws_val, 108, 12, 108, 16, style="thin")
    write_borders(ws_val, 109, 12, 109, 16, style="thin")
    
    for r in [7, 25, 37, 60, 74]:
        for col_idx in range(1, 25):
            ws_val.cell(r, col_idx).font = Font(name="Aptos Narrow", size=12, bold=True, color=WHITE)
            ws_val.cell(r, col_idx).fill = PatternFill("solid", fgColor=DARK_BG)
            
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    path = OUTPUTS / f"{ticker}_DCF.xlsx"
    wb.save(path)
    print(f"Generated {path}")
    return path

if __name__ == "__main__":
    import sys
    ticker = sys.argv[1] if len(sys.argv) > 1 else "MNDY"
    build_dcf_workbook(ticker)
